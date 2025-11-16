"""Export comparison results to Excel with comprehensive statistics.

This script generates Excel files with:
- Average prices and standard deviations
- Average computational times
- Average exercise times (1 = all exercised at maturity)
- Organized by algorithm and parameters
"""

import os
import pandas as pd
import numpy as np
from absl import app, flags
from optimal_stopping.utilities import configs_getter, read_data
from optimal_stopping.run import configs

# Telegram setup
try:
    from telegram_notifications import send_bot_message as SBM

    TELEGRAM_ENABLED = True
except:
    TELEGRAM_ENABLED = False


    class SBM:
        @staticmethod
        def send_notification(*args, **kwargs):
            pass

FLAGS = flags.FLAGS
flags.DEFINE_string("excel_output_dir", None, "Output directory for Excel files (default: latex/tables_draft)")
flags.DEFINE_string("telegram_token", "8239319342:AAGIIcoDaxJ1uauHbWfdByF4yzNYdQ5jpiA", "Telegram bot token")
flags.DEFINE_string("telegram_chat_id", "798647521", "Telegram chat ID")
flags.DEFINE_bool("send_telegram", True, "Whether to send Excel via Telegram")

ALGOS_ORDER = ["RLSM", "SRLSM", "RFQI", "SRFQI", "LSM", "DOS", "NLSM", "FQI"]


def _get_output_dir():
    """Get output directory for Excel files."""
    if FLAGS.excel_output_dir:
        return os.path.abspath(FLAGS.excel_output_dir)
    return os.path.abspath(os.path.join(
        os.path.dirname(__file__), "../../../latex/tables_draft/"))


def _format_time(seconds):
    """Convert seconds to human-readable format."""
    if pd.isna(seconds):
        return "N/A"
    hours = int(seconds // 3600)
    minutes = int((seconds % 3600) // 60)
    secs = int(seconds % 60)
    if hours > 0:
        return f"{hours}h{minutes:02d}m{secs:02d}s"
    elif minutes > 0:
        return f"{minutes}m{secs:02d}s"
    else:
        return f"{secs}s"


def extract_data_for_excel(config: configs._DefaultConfig):
    """Extract all relevant data from CSVs.

    Returns:
        DataFrame with columns:
        - All index parameters (algo, payoff, nb_stocks, etc.)
        - price_mean: Average price
        - price_std: Standard deviation of price
        - comp_time_mean: Average computational time (seconds)
        - comp_time_formatted: Human-readable comp time
        - exercise_time_mean: Average exercise time (0-1, where 1=all at maturity)
        - exercise_time_std: Std dev of exercise time
    """
    # Read ALL CSVs without filtering
    csv_paths = read_data.get_csv_paths() + read_data.get_csv_paths_draft()
    print(f"  Reading data from {len(csv_paths)} CSV files...")

    dfs = []
    for path in csv_paths:
        try:
            # Try reading with full INDEX
            df = pd.read_csv(path, index_col=read_data.INDEX)
            dfs.append(df)
        except Exception:
            # Fallback: read without index, then set index with available columns
            try:
                df = pd.read_csv(path)
                # Only use INDEX columns that exist
                available_index = [col for col in read_data.INDEX if col in df.columns]
                if available_index:
                    df = df.set_index(available_index)
                    dfs.append(df)
            except Exception as e:
                print(f"    Skipping {os.path.basename(path)}: {e}")
                pass

    if not dfs:
        raise AssertionError("No CSVs found")

    df = pd.concat(dfs)
    print(f"  Total rows read: {len(df)}")

    # Filter by payoffs
    if hasattr(config, 'payoffs') and config.payoffs:
        if 'payoff' in df.index.names:
            df = df[df.index.get_level_values('payoff').isin(config.payoffs)]
            print(f"  After payoff filter: {len(df)} rows")

    # DEBUG: Show what's actually in the data
    print(f"  Data contains:")
    for col in df.index.names[:5]:  # Just show first 5 to avoid clutter
        unique_vals = df.index.get_level_values(col).unique()
        if len(unique_vals) <= 10:
            print(f"    {col}: {list(unique_vals)}")
        else:
            print(f"    {col}: {len(unique_vals)} unique values")

    # Filter by other parameters (not algo or barriers yet)
    from optimal_stopping.utilities import filtering
    for filter_name, column_name in filtering.FILTERS:
        # Skip algo and barrier filters - handled separately below
        if filter_name in ['algos', 'barriers', 'barriers_up', 'barriers_down']:
            continue

        if column_name not in df.index.names:
            continue

        values = list(getattr(config, filter_name, []))
        if not values:
            continue

        if filter_name == "factors":
            values = [str(x) for x in values]

        rows_before = len(df)

        # Special handling for None values (match NaN, None, empty, "None")
        if None in values:
            col_values = df.index.get_level_values(column_name)
            # Match: None, NaN, empty string, or string "None"
            idx = (col_values.isna() |
                   (col_values == None) |
                   (col_values == '') |
                   (col_values == 'None') |
                   (col_values.astype(str) == 'nan'))
            # Also include other non-None values from the filter
            other_values = [v for v in values if v is not None]
            if other_values:
                idx = idx | col_values.isin(other_values)
            df = df[idx]
        else:
            df = df[df.index.get_level_values(column_name).isin(values)]

        rows_after = len(df)

        if rows_after < rows_before:
            print(f"  Filter {column_name}: {rows_before} → {rows_after} rows (config wants: {values})")

        if rows_after == 0:
            # DEBUG: Show what values are actually in the data
            print(f"  ❌ Filter {column_name} removed ALL data!")
            print(f"     Config wants: {values}")
            # Can't show actual values since df is empty, so break
            break

    print(f"  After parameter filters: {len(df)} rows")

    # NOW filter by barriers - use EXACT barrier values from config
    if 'barrier' in df.index.names and hasattr(config, 'barriers') and config.barriers:
        barrier_values = config.barriers if isinstance(config.barriers, (list, tuple)) else [config.barriers]
        # Filter to keep only rows matching the configured barrier values
        df = df[df.index.get_level_values('barrier').isin(barrier_values)]
        print(f"  Filtered by barriers {barrier_values}: {len(df)} rows")

    # Filter by algos
    if hasattr(config, 'algos') and config.algos:
        if 'algo' in df.index.names:
            df = df[df.index.get_level_values('algo').isin(config.algos)]
            print(f"  Filtered to algos {config.algos}: {len(df)} rows")

    if df.empty:
        raise AssertionError("No data after filtering")

    print(f"  ✅ Final dataset: {len(df)} rows")

    # DEBUG: Show available columns and index
    print(f"  Available columns: {list(df.columns)}")
    print(f"  Index names: {df.index.names}")



    print(f"  Sample data:")
    print(df.head(3))

    #Group by index and calculate statistics
    print(f"  Grouping by: {list(df.index.names)}")

    # DEBUG: Check if there are any duplicate index entries
    if df.index.has_duplicates:
        print(f"  ⚠️ WARNING: Index has duplicates!")
        dup_count = df.index.duplicated().sum()
        print(f"     {dup_count} duplicate index entries found")

    grouped = df.groupby(level=list(range(len(df.index.names))))  # Use level numbers instead of names
    print(f"  Number of groups: {len(grouped)}")

    # DEBUG: Try to see a sample group
    try:
        first_group_key = list(grouped.groups.keys())[0]
        first_group = grouped.get_group(first_group_key)
        print(f"  Sample group size: {len(first_group)} rows")
        print(f"  Sample group columns: {list(first_group.columns)}")
    except Exception as e:
        print(f"  ⚠️ Could not get sample group: {e}")

    stats_dict = {}

    # Add price stats if available
    if 'price' in df.columns:
        try:
            price_mean = grouped['price'].mean()
            print(f"  Price mean calculated: {len(price_mean)} entries")
            if len(price_mean) == 0:
                print(f"  ⚠️ WARNING: price_mean is empty!")
                print(f"     Trying alternative aggregation method...")
                # Try using agg() instead
                price_stats = grouped['price'].agg(['mean', 'std'])
                print(f"     Alternative method result: {len(price_stats)} rows")
                if len(price_stats) > 0:
                    stats_dict['price_mean'] = price_stats['mean']
                    stats_dict['price_std'] = price_stats['std'].fillna(0)
                    print(f"  ✅ Added price statistics (via alternative method)")
            else:
                stats_dict['price_mean'] = price_mean
                stats_dict['price_std'] = grouped['price'].std().fillna(0)
                print(f"  ✅ Added price statistics")
        except Exception as e:
            print(f"  ❌ Error calculating price stats: {e}")
            import traceback
            traceback.print_exc()
    else:
        print(f"  ⚠️ No 'price' column found")

    # Add comp_time if available
    if 'comp_time' in df.columns:
        stats_dict['comp_time_mean'] = grouped['comp_time'].mean()
        stats_dict['comp_time_std'] = grouped['comp_time'].std().fillna(0)
        print(f"  ✅ Added comp_time statistics")

    # Add duration if available
    if 'duration' in df.columns:
        stats_dict['duration_mean'] = grouped['duration'].mean()
        stats_dict['duration_std'] = grouped['duration'].std().fillna(0)
        print(f"  ✅ Added duration statistics")

    # Add exercise_time if available (average time of exercise, normalized 0-1)
    if 'exercise_time' in df.columns:
        stats_dict['exercise_time_mean'] = grouped['exercise_time'].mean()
        stats_dict['exercise_time_std'] = grouped['exercise_time'].std().fillna(0)
        print(f"  ✅ Added exercise_time statistics")

    if not stats_dict:
        print(f"  ❌ No statistics could be calculated - no recognized columns")
        raise AssertionError("No statistics columns found")

    # Create result DataFrame
    result_df = pd.DataFrame(stats_dict)
    print(f"  Created stats DataFrame: {len(result_df)} rows")

    # Add formatted time column
    if 'comp_time_mean' in result_df.columns:
        result_df['comp_time_formatted'] = result_df['comp_time_mean'].apply(_format_time)
    elif 'duration_mean' in result_df.columns:
        result_df['comp_time_formatted'] = result_df['duration_mean'].apply(_format_time)

    # Reset index to make parameters into columns
    result_df = result_df.reset_index()
    print(f"  Final result: {len(result_df)} rows")

    return result_df


def create_excel_workbook(label: str, config: configs._DefaultConfig):
    """Create Excel workbook with multiple sheets for different views.

    Sheets:
    1. Summary: All data in wide format (algos as columns)
    2. Prices: Just prices with std
    3. Times: Computational times
    4. Exercise: Exercise time statistics
    5. Raw: Long format with all data
    """
    print(f"\n📊 Processing {label}...")

    # Extract data
    df = extract_data_for_excel(config)

    if df.empty:
        print(f"  ⚠️ No data for {label}")
        return None

    print(f"  ✅ Extracted {len(df)} rows")

    # Get output path
    output_dir = _get_output_dir()
    os.makedirs(output_dir, exist_ok=True)
    excel_path = os.path.join(output_dir, f"{label}_comprehensive.xlsx")

    # Get varying parameters (exclude algo and single-value params)
    index_cols = [col for col in df.columns if col in read_data.INDEX and col != 'algo']
    varying_params = [col for col in index_cols if df[col].nunique() > 1]

    print(f"  Varying parameters: {varying_params}")

    # Create Excel writer
    try:
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            # Sheet 1: RAW DATA (long format)
            df.to_excel(writer, sheet_name='Raw Data', index=False)

            # Sheet 2: PRICES (wide format - algos as columns)
            if 'algo' in df.columns:
                prices_df = create_wide_format_sheet(
                    df, varying_params, 'algo',
                    value_cols=['price_mean', 'price_std'],
                    sheet_name='Prices'
                )
                if prices_df is not None:
                    prices_df.to_excel(writer, sheet_name='Prices', index=False)

            # Sheet 3: COMPUTATIONAL TIMES
            if 'comp_time_mean' in df.columns or 'duration_mean' in df.columns:
                time_col = 'comp_time_mean' if 'comp_time_mean' in df.columns else 'duration_mean'
                time_std_col = 'comp_time_std' if 'comp_time_std' in df.columns else 'duration_std'

                times_df = create_wide_format_sheet(
                    df, varying_params, 'algo',
                    value_cols=[time_col, time_std_col, 'comp_time_formatted'],
                    sheet_name='Comp Times'
                )
                if times_df is not None:
                    times_df.to_excel(writer, sheet_name='Comp Times', index=False)

            # Sheet 4: EXERCISE TIMES
            if 'exercise_time_mean' in df.columns:
                exercise_df = create_wide_format_sheet(
                    df, varying_params, 'algo',
                    value_cols=['exercise_time_mean', 'exercise_time_std'],
                    sheet_name='Exercise Times'
                )
                if exercise_df is not None:
                    exercise_df.to_excel(writer, sheet_name='Exercise Times', index=False)

            # Sheet 5: SUMMARY (prices + times in compact format)
            summary_df = create_summary_sheet(df, varying_params)
            if summary_df is not None:
                summary_df.to_excel(writer, sheet_name='Summary', index=False)

            # Format workbook
            format_workbook(writer)

        print(f"  ✅ Excel file: {excel_path}")

        # Send via Telegram
        if TELEGRAM_ENABLED and FLAGS.send_telegram:
            try:
                SBM.send_notification(
                    token=FLAGS.telegram_token,
                    text=f"📊 Excel: {label}\n\nSheets: Raw, Prices, Times, Exercise, Summary",
                    files=[excel_path],
                    chat_id=FLAGS.telegram_chat_id
                )
                print(f"  ✅ Sent via Telegram")
            except Exception as e:
                print(f"  ⚠️ Telegram failed: {e}")

        return excel_path

    except Exception as e:
        print(f"  ❌ Failed to create Excel: {e}")
        import traceback
        traceback.print_exc()
        return None


def create_wide_format_sheet(df, varying_params, pivot_col, value_cols, sheet_name):
    """Create wide format with algorithms as columns.

    Args:
        df: Source DataFrame
        varying_params: List of parameter columns
        pivot_col: Column to pivot (usually 'algo')
        value_cols: List of value columns to include
        sheet_name: Name for debugging
    """
    if pivot_col not in df.columns:
        print(f"  ⚠️ {sheet_name}: No {pivot_col} column")
        return None

    # Filter to relevant columns
    cols_to_keep = varying_params + [pivot_col] + [col for col in value_cols if col in df.columns]
    df_filtered = df[cols_to_keep].copy()

    if len(varying_params) == 0:
        # No varying params - just pivot
        result = df_filtered.pivot(columns=pivot_col, values=[col for col in value_cols if col in df.columns])
    else:
        # Pivot with index
        result = df_filtered.pivot_table(
            index=varying_params,
            columns=pivot_col,
            values=[col for col in value_cols if col in df.columns],
            aggfunc='first'  # Should only be one value per combo
        )

    # Flatten column names
    if isinstance(result.columns, pd.MultiIndex):
        result.columns = [f"{col[0]}_{col[1]}" if len(col) > 1 else str(col[0])
                          for col in result.columns]

    # Reorder algorithm columns
    if pivot_col == 'algo':
        result = reorder_algo_columns(result)

    result = result.reset_index()

    print(f"  {sheet_name}: {result.shape}")
    return result


def create_summary_sheet(df, varying_params):
    """Create summary sheet with prices (mean±std) and comp times."""
    if 'algo' not in df.columns:
        return None

    algos = sorted([a for a in df['algo'].unique() if a in ALGOS_ORDER],
                   key=lambda x: ALGOS_ORDER.index(x))

    # Create compact format: "mean ± std"
    summary_data = []

    for _, group in df.groupby(varying_params if varying_params else [df.index]):
        row = {}

        # Add varying parameters
        if varying_params:
            for param in varying_params:
                row[param] = group[param].iloc[0]

        # Add data for each algorithm
        for algo in algos:
            algo_data = group[group['algo'] == algo]
            if len(algo_data) == 0:
                row[f"{algo}_price"] = "N/A"
                row[f"{algo}_time"] = "N/A"
                continue

            # Price: mean ± std
            mean = algo_data['price_mean'].iloc[0]
            std = algo_data['price_std'].iloc[0]
            row[f"{algo}_price"] = f"{mean:.2f} ± {std:.2f}"

            # Time
            if 'comp_time_formatted' in algo_data.columns:
                row[f"{algo}_time"] = algo_data['comp_time_formatted'].iloc[0]
            else:
                row[f"{algo}_time"] = "N/A"

        summary_data.append(row)

    if not summary_data:
        return None

    result = pd.DataFrame(summary_data)
    print(f"  Summary: {result.shape}")
    return result


def reorder_algo_columns(df):
    """Reorder columns to put algorithms in ALGOS_ORDER."""
    cols = list(df.columns)

    # Separate algo columns from others
    algo_cols = []
    other_cols = []

    for col in cols:
        is_algo_col = False
        for algo in ALGOS_ORDER:
            if algo in col:
                algo_cols.append(col)
                is_algo_col = True
                break
        if not is_algo_col:
            other_cols.append(col)

    # Sort algo columns by ALGOS_ORDER
    def algo_sort_key(col):
        for i, algo in enumerate(ALGOS_ORDER):
            if algo in col:
                return (i, col)
        return (999, col)

    algo_cols_sorted = sorted(algo_cols, key=algo_sort_key)

    # Recombine
    return df[other_cols + algo_cols_sorted]


def format_workbook(writer):
    """Apply formatting to Excel workbook."""
    try:
        from openpyxl.styles import Font, Alignment, PatternFill

        for sheet_name in writer.sheets:
            ws = writer.sheets[sheet_name]

            # Header formatting
            for cell in ws[1]:
                cell.font = Font(bold=True, size=11)
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.font = Font(bold=True, size=11, color="FFFFFF")
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter

                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass

                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            # Freeze header row
            ws.freeze_panes = 'A2'

        print("  ✅ Formatting applied")
    except Exception as e:
        print(f"  ⚠️ Formatting failed: {e}")


def write_all_excel():
    """Generate Excel files for all configs."""
    generated_files = []

    for config_name, config in configs_getter.get_configs():
        representations = list(config.representations)

        # Only process table representations
        if not any(rep in ["TablePrice", "TableDuration", "TablePriceDuration"]
                   for rep in representations):
            print(f"⏭️  Skipping {config_name} (no table representation)")
            continue

        try:
            excel_path = create_excel_workbook(config_name, config)
            if excel_path:
                generated_files.append(excel_path)
        except Exception as e:
            print(f"❌ Error processing {config_name}: {e}")
            import traceback
            traceback.print_exc()

    return generated_files


def main(argv):
    del argv

    print("\n" + "=" * 70)
    print("GENERATING EXCEL EXPORTS")
    print("=" * 70)

    try:
        # Send start notification
        if TELEGRAM_ENABLED and FLAGS.send_telegram:
            SBM.send_notification(
                token=FLAGS.telegram_token,
                text='📊 Starting Excel generation...',
                chat_id=FLAGS.telegram_chat_id
            )

        # Generate Excel files
        generated_files = write_all_excel()

        print("\n" + "=" * 70)
        print(f"✅ Generated {len(generated_files)} Excel file(s)")
        for f in generated_files:
            print(f"  - {os.path.basename(f)}")
        print("=" * 70 + "\n")

        # Send completion notification
        if TELEGRAM_ENABLED and FLAGS.send_telegram:
            SBM.send_notification(
                token=FLAGS.telegram_token,
                text=f'✅ Excel generation complete!\n\n{len(generated_files)} file(s) created',
                chat_id=FLAGS.telegram_chat_id
            )

    except Exception as e:
        print(f"\n❌ FATAL ERROR: {e}")
        import traceback
        traceback.print_exc()

        if TELEGRAM_ENABLED and FLAGS.send_telegram:
            SBM.send_notification(
                token=FLAGS.telegram_token,
                text=f'❌ Excel generation failed:\n{e}',
                chat_id=FLAGS.telegram_chat_id
            )
        raise


if __name__ == "__main__":
    app.run(main)